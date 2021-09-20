import pytumblr
import openpyxl
import time

pathxls = ".\\Cats.xlsx"
pathcat = ".\\cats\\"
wb = openpyxl.load_workbook(pathxls)
sheet = wb['cats']
max_row = sheet.max_row
numcat = sheet['D1'].value
cantidad = 0
tope = 93

# Authenticate in Tumblr via OAuth
client = pytumblr.TumblrRestClient(
  'OAuth Key Here'
)

# Make the request

while cantidad < tope:

    client.create_photo("catseverydays", state="queue", tags=["cats", "facts", "kitties", "catfacts"], caption=str(sheet.cell(row=numcat, column=1).value),
                        tweet=str(sheet.cell(row=numcat, column=1).value),
                        data=pathcat+str(numcat)+".jpg")

    print (client.queue("catseverydays"))
    
    print ("Cat number: " + str (numcat))
    
    numcat += 1
    cantidad += 1
    sheet['D1'] = numcat
    time.sleep(2)
    wb.save(pathxls)
    
print ("Terminado")